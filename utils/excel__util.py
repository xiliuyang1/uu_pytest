# 1. 导入依赖
import openpyxl
from openpyxl.utils import get_column_letter


# 2. 打开Excel文件
def open_excel_file(file_name):
    workbook = openpyxl.load_workbook(file_name)
    return workbook


# 3. 读取Sheet
def read_excel_sheet(workbook, sheet_name):
    sheet = workbook[sheet_name]
    return sheet


# 4. 获取数据行数和列数
def get_row_col_count(sheet):
    row_count = sheet.max_row
    col_count = sheet.max_column
    return row_count, col_count


# 5. 获取表头名称
def get_header_names(sheet, col_count):
    header_list = []
    for col in range(1, col_count + 1):
        header_list.append(sheet.cell(row=1, column=col).value)
    return header_list


# 6. 读取数据
def read_data_rows(sheet, row_count, col_count):
    data = []
    for row in range(2, row_count + 1):
        row_values = {}
        for col in range(1, col_count + 1):
            value = sheet.cell(row=row, column=col).value
            row_values[get_column_letter(col)] = value
        # 在读取 Excel 表格时，可以通过该函数将每一列的数字转换成对应的英文字母作为 dictionary 的 key，方便后续数据的处理。
        data.append(row_values)
    return data


if __name__ == '__main__':
    print(1198469-1213469)